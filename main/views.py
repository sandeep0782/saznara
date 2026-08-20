import os
import re
from copy import copy
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import xlrd
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from xlutils.copy import copy as xl_copy

from accounts.decorators import admin_only
from main.forms import *
from main.marketplaces.flipkart.validator import validate_flipkart_template
from main.marketplaces.meesho.validator import *
from main.marketplaces.myntra.mappings import *
from main.marketplaces.myntra.validator import (
    validate_myntra_template,
)
from main.marketplaces.snapdeal.validator import validate_snapdeal_template
from main.models import SKU, Article_Type, Brand, Color, Gender, Size, Unit
from main.services.marketplace_mapping import get_marketplace_value
from main.template.mappings import *

from .models import *


# Create your views here.
@login_required
def home(request):

    return render(request, "index.html")


@admin_only
@login_required
def View__Vendors(request):

    search_query = request.GET.get("search", "").strip()

    # Admin sees all profiles
    vendor_list = Profile.objects.all().order_by("-id")

    if search_query:
        vendor_list = vendor_list.filter(company__icontains=search_query)

    paginator = Paginator(vendor_list, 10)

    page = request.GET.get("page")

    try:
        vendors = paginator.page(page)

    except PageNotAnInteger:
        vendors = paginator.page(1)

    except EmptyPage:
        vendors = paginator.page(paginator.num_pages)

    return render(
        request,
        "bag/vendor/view_vendor.html",
        {
            "vendors": vendors,
            "search_query": search_query,
        },
    )


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404


@login_required
@admin_only
def Change__Vendor(request, id):

    vendor = get_object_or_404(Profile, id=id)

    if request.method == "POST":
        vendor.user.first_name = request.POST.get("first_name")

        vendor.user.last_name = request.POST.get("last_name")

        vendor.user.email = request.POST.get("email")

        vendor.user.save()

        vendor.company = request.POST.get("company")

        vendor.mobile_no = request.POST.get("mobile_no")

        vendor.address = request.POST.get("address")

        vendor.pin = request.POST.get("pin")

        if request.FILES.get("image"):
            vendor.image = request.FILES["image"]

        vendor.save()

        messages.success(request, "Vendor updated successfully.")

        return redirect("view_vendors")

    return render(request, "bag/vendor/change_vendor.html", {"vendor": vendor})


@login_required
def View__Brand(request):
    search_query = request.GET.get("search", "").strip()

    if request.user.is_superuser:
        brand_list = Brand.objects.all().order_by("-id")
    else:
        brand_list = Brand.objects.filter(vendor=request.user.profile).order_by("-id")

    if search_query:
        brand_list = brand_list.filter(name__icontains=search_query)

    paginator = Paginator(brand_list, 10)
    page = request.GET.get("page")

    try:
        brand = paginator.page(page)
    except PageNotAnInteger:
        brand = paginator.page(1)
    except EmptyPage:
        brand = paginator.page(paginator.num_pages)

    return render(
        request,
        "bag/brand/view_brand.html",
        {
            "brand": brand,
            "search_query": search_query,
        },
    )


@login_required
def Change__Brand(request, id=None):
    brand = None
    if id:
        try:
            if request.user.is_superuser:
                brand = Brand.objects.get(id=id)
            else:
                brand = Brand.objects.get(id=id, vendor=request.user.profile)
        except Brand.DoesNotExist:
            messages.info(request, "Brand does not exist")
            return redirect("view_brand")
    if request.method == "POST":
        form = BrandForm(request.POST, request.FILES, instance=brand)
        if form.is_valid():
            pro = form.save(commit=False)

            # Assign vendor
            if not request.user.is_superuser:
                pro.vendor = request.user.profile

                # Prevent multiple brands per vendor
                if (
                    not id
                    and Brand.objects.filter(vendor=request.user.profile).exists()
                ):
                    messages.error(request, "You have already created your brand.")
                    return redirect("view_brand")

            if id:
                pro.modified_by = request.user
                messages.info(request, "Brand modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Brand added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_brand")
        else:
            messages.info(request, form.errors)
    d = {"brand": brand}
    return render(request, "bag/brand/change_brand.html", d)


@login_required
def Delete__Brand(request, id):
    if not request.user.is_superuser:
        messages.error(request, "Only the administrator can delete brands.")
        return redirect("view_brand")

    try:
        brand = Brand.objects.get(id=id)
        brand.delete()
        messages.success(request, "Brand deleted successfully.")
    except Brand.DoesNotExist:
        messages.error(request, "Brand not found.")

    return redirect("view_brand")


@login_required
def View__Article__Type(request):
    search_query = request.GET.get("search", "").strip()
    at_list = Article_Type.objects.all().order_by("-id")
    if search_query:
        at_list = at_list.filter(name=search_query)

    paginator = Paginator(at_list, 10)  # Show 10 products per page
    page = request.GET.get("page")
    try:
        at = paginator.page(page)
    except PageNotAnInteger:
        at = paginator.page(1)
    except EmptyPage:
        at = paginator.page(paginator.num_pages)
    d = {
        "at": at,
        "search_query": search_query,
    }  # Pass the search query to the template
    return render(request, "bag/at/view_article_type.html", d)


@staff_member_required
def Change__Article__Type(request, id=None):
    at = None
    if id:
        try:
            at = Article_Type.objects.get(id=id)
        except Article_Type.DoesNotExist:
            messages.info(request, "Article Type does not exist")
            return redirect("view_article_type")

    if request.method == "POST":
        form = ATForm(request.POST, request.FILES, instance=at)
        if form.is_valid():
            pro = form.save(commit=False)  # Don't save yet to modify additional fields
            if id:
                pro.modified_by = request.user  # Set the user who modified the SKU
                messages.info(request, "Article Type modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Article Type added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_article_type")
        else:
            messages.info(request, form.errors)
    d = {"at": at}
    return render(request, "bag/at/change_article_type.html", d)


@staff_member_required
def Delete__Article__Type(request, id):
    at = Article_Type.objects.filter(id=id)
    at.delete()
    return redirect("view_article_type")


@login_required
def View__Gender(request):
    gender = Gender.objects.all().order_by("-id")
    d = {"gender": gender}  # Pass the search query to the template
    return render(request, "bag/gender/view_gender.html", d)


@staff_member_required
def Change__Gender(request, id=None):
    gender = None
    if id:
        try:
            gender = Gender.objects.get(id=id)
        except Gender.DoesNotExist:
            messages.info(request, "Gender Type does not exist")
            return redirect("view_gender")

    if request.method == "POST":
        form = GenderForm(request.POST, request.FILES, instance=gender)
        if form.is_valid():
            pro = form.save(commit=False)  # Don't save yet to modify additional fields
            if id:
                pro.modified_by = request.user  # Set the user who modified the SKU
                messages.info(request, "Gender modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Gender added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_gender")
        else:
            messages.info(request, form.errors)
    d = {"gender": gender}
    return render(request, "bag/gender/change_gender.html", d)


@staff_member_required
def Delete__Gender(request, id):
    gender = Gender.objects.filter(id=id)
    gender.delete()
    return redirect("view_gender")


@login_required
def View__Size(request):
    search_query = request.GET.get("search", "").strip()
    size_list = Size.objects.all().order_by("-id")
    if search_query:
        size_list = size_list.filter(Q(size=search_query) | Q(abb=search_query))

    paginator = Paginator(size_list, 10)  # Show 10 products per page
    page = request.GET.get("page")
    try:
        size = paginator.page(page)
    except PageNotAnInteger:
        size = paginator.page(1)
    except EmptyPage:
        size = paginator.page(paginator.num_pages)
    d = {
        "size": size,
        "search_query": search_query,
    }  # Pass the search query to the template
    return render(request, "bag/size/view_size.html", d)


@staff_member_required
def Change__Size(request, id=None):
    size = None
    if id:
        try:
            size = Size.objects.get(id=id)
        except Size.DoesNotExist:
            messages.info(request, "Size Type does not exist")
            return redirect("view_size")

    if request.method == "POST":
        form = SizeForm(request.POST, request.FILES, instance=size)
        if form.is_valid():
            pro = form.save(commit=False)  # Don't save yet to modify additional fields
            if id:
                pro.modified_by = request.user  # Set the user who modified the SKU
                messages.info(request, "Size modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Size added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_size")
        else:
            messages.info(request, form.errors)
    d = {"size": size}
    return render(request, "bag/size/change_size.html", d)


@staff_member_required
def Delete__Size(request, id):
    size = Size.objects.filter(id=id)
    size.delete()
    return redirect("view_size")


@login_required
def View__UOM(request):
    uom = Unit.objects.all().order_by("-id")
    d = {"uom": uom}  # Pass the search query to the template
    return render(request, "bag/uom/view_uom.html", d)


@staff_member_required
def Change__UOM(request, id=None):
    uom = None
    if id:
        try:
            uom = Unit.objects.get(id=id)
        except Unit.DoesNotExist:
            messages.info(request, "Unit Type does not exist")
            return redirect("view_uom")

    if request.method == "POST":
        form = UnitForm(request.POST, request.FILES, instance=uom)
        if form.is_valid():
            pro = form.save(commit=False)  # Don't save yet to modify additional fields
            if id:
                pro.modified_by = request.user  # Set the user who modified the SKU
                messages.info(request, "Unit modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Unit added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_uom")
        else:
            messages.info(request, form.errors)
    d = {"uom": uom}
    return render(request, "bag/uom/change_uom.html", d)


@staff_member_required
def Delete__UOM(request, id):
    uom = Unit.objects.filter(id=id)
    uom.delete()
    return redirect("view_uom")


@login_required
def View__Color(request):
    search_query = request.GET.get("search", "").strip()
    color_list = Color.objects.all().order_by("-id")
    if search_query:
        color_list = color_list.filter(color=search_query)

    paginator = Paginator(color_list, 10)  # Show 10 products per page
    page = request.GET.get("page")
    try:
        color = paginator.page(page)
    except PageNotAnInteger:
        color = paginator.page(1)
    except EmptyPage:
        color = paginator.page(paginator.num_pages)
    d = {
        "color": color,
        "search_query": search_query,
    }  # Pass the search query to the template
    return render(request, "bag/color/view_color.html", d)


@staff_member_required
def Change__Color(request, id=None):
    color = None
    if id:
        try:
            color = Color.objects.get(id=id)
        except Color.DoesNotExist:
            messages.info(request, "Color Type does not exist")
            return redirect("view_color")

    if request.method == "POST":
        form = ColorForm(request.POST, request.FILES, instance=color)
        if form.is_valid():
            pro = form.save(commit=False)  # Don't save yet to modify additional fields
            if id:
                pro.modified_by = request.user  # Set the user who modified the SKU
                messages.info(request, "Color modified successfully")
            else:
                pro.created_by = request.user  # Set the user who created the SKU
                messages.info(request, "Color added successfully")
            pro.save()  # Now save the SKU with all fields
            return redirect("view_color")
        else:
            messages.info(request, form.errors)
    d = {"color": color}
    return render(request, "bag/color/change_color.html", d)


@staff_member_required
def Delete__Color(request, id):
    c = Color.objects.filter(id=id)
    c.delete()
    return redirect("view_color")


@login_required
def get_filtered_skus(request):
    search_query = request.GET.get("search", "").strip()

    sku_list = SKU.objects.all().order_by("-id")

    if search_query:
        search_terms = search_query.split("%")

        for term in search_terms:
            term = term.strip()
            if term:
                sku_list = sku_list.filter(
                    Q(ref_no__icontains=term)
                    | Q(vendor__company__icontains=term)
                    | Q(sku__icontains=term)
                    | Q(brand__name__icontains=term)
                    | Q(style_no__icontains=term)
                    | Q(color__color__icontains=term)
                    | Q(article_type__name__icontains=term)
                )

    return sku_list, search_query


@login_required(login_url="login")
def View__SKU(request):

    # GET PARAMETERS
    search_query = request.GET.get("search", "").strip()
    export = request.GET.get("export")

    # BASE QUERYSET
    if request.user.is_superuser:
        # Admin can see all SKUs
        sku_list = SKU.objects.all().order_by("-id")
    else:
        # Vendor can see only own SKUs
        sku_list = SKU.objects.filter(vendor=request.user.profile).order_by("-id")

    # FILTERING
    if search_query:
        search_terms = search_query.split("%")

        for term in search_terms:
            term = term.strip()

            if term:
                sku_list = sku_list.filter(
                    Q(ref_no__icontains=term)
                    | Q(vendor__company__icontains=term)
                    | Q(sku__icontains=term)
                    | Q(brand__name__icontains=term)
                    | Q(style_no__icontains=term)
                    | Q(color__color__icontains=term)
                    | Q(article_type__name__icontains=term)
                )

    # EXPORT
    if export == "meesho":
        return Meesho_Template(request, sku_list)

    if export == "flipkart":
        return Flipkart_Template(request, sku_list)

    if export == "snapdeal":
        return Snapdeal_Template(request, sku_list)

    if export == "myntra":
        return Myntra_Template(request, sku_list)

    # PAGINATION
    paginator = Paginator(sku_list, 10)
    page = request.GET.get("page")

    try:
        sku = paginator.page(page)
    except PageNotAnInteger:
        sku = paginator.page(1)
    except EmptyPage:
        sku = paginator.page(paginator.num_pages)

    return render(
        request, "sku/view_sku.html", {"sku": sku, "search_query": search_query}
    )


@login_required
def Change__SKU(request, pid=None):
    sku = None

    brand = (
        Brand.objects.all()
        if request.user.is_superuser
        else Brand.objects.filter(vendor=request.user.profile)
    )
    gender = Gender.objects.all()
    article_type = Article_Type.objects.all()
    size = Size.objects.all()
    color = Color.objects.all()
    vendor = Profile.objects.all()

    selected_gender = Gender.objects.filter(name__iexact="Women").first()
    selected_article_type = Article_Type.objects.filter(name__iexact="Sarees").first()
    selected_size = Size.objects.filter(size__iexact="Free Size").first()
    if pid:
        sku = SKU.objects.filter(id=pid).first()
        if not sku:
            messages.error(request, "SKU does not exist")
            return redirect("view_sku")

    if request.method == "POST":
        form = SKUForm(request.POST, request.FILES, instance=sku)

        if form.is_valid():
            obj = form.save(commit=False)

            if pid:
                obj.modified_by = request.user
                messages.success(request, "SKU updated successfully")
            else:
                obj.created_by = request.user
                obj.vendor = request.user.profile
                obj.gender = selected_gender
                obj.article_type = selected_article_type
                obj.size = selected_size
                messages.success(request, "SKU created successfully")

            obj.save()
            return redirect("view_sku")

        else:
            # IMPORTANT: return form WITH ERRORS
            messages.error(request, "Please correct the errors below")
    else:
        form = SKUForm(instance=sku)

    context = {
        "form": form,
        "sku": sku,
        "brand": brand,
        "gender": gender,
        "article_type": article_type,
        "size": size,
        "color": color,
        "vendor": vendor,
        "BLOUSE_CHOICES": BLOUSE_CHOICES,
        "BORDER_CHOICES": BORDER_CHOICES,
        "PRINT_PATTERN_TYPE_CHOICES": PRINT_PATTERN_TYPE_CHOICES,
        "SAREE_FABRIC_CHOICES": SAREE_FABRIC_CHOICES,
        "TRANSPARENCY_CHOICES": TRANSPARENCY_CHOICES,
        "TYPE_CHOICES": TYPE_CHOICES,
        "BLOUSE_LENGTH_SIZE_CHOICES": BLOUSE_LENGTH_SIZE_CHOICES,
        "SAREE_LENGTH_SIZE_CHOICES": SAREE_LENGTH_SIZE_CHOICES,
        "BLOUSE_COLOR_CHOICES": BLOUSE_COLOR_CHOICES,
        "BLOUSE_FABRIC_CHOICES": BLOUSE_FABRIC_CHOICES,
        "BLOUSE_PATTERN_CHOICES": BLOUSE_PATTERN_CHOICES,
        "BORDER_WIDTH_CHOICES": BORDER_WIDTH_CHOICES,
        "LOOM_TYPE_CHOICES": LOOM_TYPE_CHOICES,
        "OCCASION_CHOICES": OCCASION_CHOICES,
        "ORNAMENTATION_CHOICES": ORNAMENTATION_CHOICES,
        "PALLU_DETAILS_CHOICES": PALLU_DETAILS_CHOICES,
        "DESIGN_PATTERN_CHOICES": DESIGN_PATTERN_CHOICES,
    }

    return render(request, "sku/change_sku.html", context)


@login_required
def Delete__SKU(request, pid):
    sku = SKU.objects.get(id=pid)  # Fetch the SKU object
    sku.delete()
    return redirect("view_sku")


@login_required
def Print__SKU(request, pid):
    sku = SKU.objects.get(id=pid)  # Fetch the SKU object
    if len(sku.sku) < 18:
        messages.warning(
            request,
            "SKU code is not in a proper format, hence barcode cannot be printed",
        )
        return redirect("view_sku")

    if sku.brand.name == "SUHA":
        return render(request, "sku/print_suha.html", {"sku": sku})
    elif (
        sku.brand.name == "NYRIKA"
        or sku.brand.name == "INDIE PICKS"
        or sku.brand.name == "FYREROSE"
        or sku.brand.name == "BUDA JEANS"
        or sku.brand.name == "SVARAA"
    ):
        return render(request, "sku/print_ajio.html", {"sku": sku})
    else:
        return render(request, "sku/print_myntra.html", {"sku": sku})


@login_required
def Print__Barcode(request, pid):
    sku = SKU.objects.get(id=pid)  # Fetch the SKU object
    if len(sku.sku) < 18:
        messages.warning(
            request,
            "SKU code is not in a proper format, hence barcode cannot be printed",
        )
        return redirect("view_sku")
    return render(request, "sku/print_barcode.html", {"sku": sku})


@login_required
def Copy__SKU(request, pid=None):

    if not pid:
        messages.error(request, "No SKU ID provided.")
        return redirect("view_sku")

    try:
        sku = SKU.objects.get(id=pid)
    except SKU.DoesNotExist:
        messages.error(request, "SKU does not exist.")
        return redirect("view_sku")

    with transaction.atomic():
        # Get all SKUs of the same vendor
        skus = SKU.objects.filter(vendor=sku.vendor)

        max_number = 0
        prefix = ""
        width = 4

        for s in skus:
            style = str(s.style_no or "")
            match = re.search(r"([A-Za-z]*)(\d+)$", style)

            if match:
                prefix = match.group(1)
                number = int(match.group(2))
                width = len(match.group(2))

                max_number = max(max_number, number)

        # Generate next style number
        new_style_no = f"{prefix}{str(max_number + 1).zfill(width)}"

        # Clone the SKU
        new_sku = copy(sku)
        new_sku.pk = None
        new_sku.id = None

        # Set new style number
        new_sku.style_no = new_style_no

        # Save (your save() method will generate SKU, VAN, EAN, Barcode)
        new_sku.save()

    messages.success(request, f"SKU copied successfully with Style No {new_style_no}")

    return redirect("view_sku")


@login_required
def Snapdeal_Template1(request, sku_list):
    sku_list, _ = get_filtered_skus(request)

    validation_errors = validate_snapdeal_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/snapdeal_error.html",
            {"validation_errors": validation_errors},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snapdeal Saree Template"

    # =========================================================
    # SNAPDEAL HEADERS
    # =========================================================
    headers = [
        "Offer Group Name",
        "SKU Code",
        "Brand",
        "Product Name",
        "Color",
        "Fabric",
        "Set Contents",
        "Type",
        "Manufacturer's Name & Address",
        "Saree Length(in metre)",
        "Blouse Piece Length (in meter)",
        "Country of Origin or Manufacture or Assembly",
        "Packer's Name & Address",
        "Net Contents",
        "Pattern",
        "Pack",
        "Saree Type",
        "Style Code/Name",
        "MRP",
        "Selling Price",
        "Inventory",
        "Shipping Time in Days",
        "Height (cm)",
        "Width (cm)",
        "Length (cm)",
        "Weight (g)",
        "Image 1",
        "Image 2",
        "Image 3",
        "Image 4",
        "Image 5",
        "Image 6",
        "Image 7",
        "Image 8",
        "Image 9",
        "Image 10",
        "Image 11",
        "Image 12",
        "Description",
        "EAN",
        "UPC",
        "Blouse Fabric",
        "Blouse Color",
        "Border Specific",
        "Saree width(in metre)",
        "Product Weight (in kg)",
        "Common or Generic Name of the commodity",
        "Importer's Name & Address",
        "Marketer's Name & Address",
        "Blouse Pattern",
        "Pattern or Print Type",
        "Brand Color",
        "Generic Keywords",
    ]

    ws.append(headers)

    # =========================================================
    # ROWS
    # =========================================================
    for sku in sku_list:
        vendor_address = ""

        if sku.vendor:
            vendor_address = f"{sku.vendor.company}, {sku.vendor.address}"

        color = ""
        if sku.color:
            color = SNAPDEAL_COLOR_MAPPING.get(sku.color.color, sku.color.color)

        blouse_color = ""
        if sku.get_blouse_color_display:
            blouse_color = MEESHO_COLOR_MAPPING.get(
                sku.get_blouse_color_display(), sku.get_blouse_color_display()
            )

        occasion = ""
        if sku.occasion:
            occasion = MEESHO_OCCASION_MAPPING.get(
                sku.get_occasion_display(), sku.get_occasion_display()
            )

        ws.append(
            [
                # Offer Group Name
                "",
                # SKU Code
                sku.sku or "",
                # Brand
                sku.brand.name if sku.brand else "",
                # Product Name
                "",  # Is should be blank as per snapdeal template
                # Color
                color,
                # Fabric
                sku.get_saree_fabric_display() if sku.saree_fabric else "",
                # Set Contents
                SNAPDEAL_SET_CONTENTS_MAPPING.get(sku.blouse, "Without Blouse Piece"),
                # Type
                sku.get_type_display() if sku.type else "",
                # Manufacturer Name & Address
                vendor_address,
                # Saree Length
                float(sku.saree_length or 5.5),
                # Blouse Length
                float(sku.blouse_length or 0.8),
                # Country
                "India",
                # Packer
                vendor_address,
                # Net Contents
                "1",
                # Pattern
                sku.get_pattern_display() if sku.pattern else "",
                # Pack
                "Pack of 1",
                # Saree Type
                "Regular Saree",
                # Style Code/Name
                sku.sku or sku.ref_no or "",
                # MRP
                sku.mrp or "",
                # Selling Price
                (sku.sale_price + 300) if sku.sale_price else "",
                # Inventory
                100,
                # Shipping Time
                "2",
                # Height
                "30",
                # Width
                "25",
                # Length
                "5",
                # Weight grams
                400,
                # Images
                sku.product_image_link_1 or "",
                sku.product_image_link_2 or "",
                sku.product_image_link_3 or "",
                sku.product_image_link_4 or "",
                # Image 5-12 blank
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                # Description
                sku.style_description or "",
                # EAN
                "",
                # UPC
                "",
                # Blouse Fabric
                sku.get_blouse_fabric_display() if sku.blouse_fabric else "",
                # Blouse Color
                # blouse_color,
                get_snapdeal_blouse_color(sku.blouse_color) if sku.blouse_color else "",
                # Border Specific
                sku.get_border_display() if sku.border else "",
                # Saree width
                "1.08",
                # Product Weight kg
                0.4,
                # Generic Name
                "Saree",
                # Importer
                "",
                # Marketer
                vendor_address,
                # Blouse Pattern
                get_snapdeal_blouse_pattern(sku.blouse_pattern)
                if sku.blouse_pattern
                else "",
                # sku.get_blouse_pattern_display()
                # if sku.blouse_pattern else "",
                # Pattern or Print Type
                sku.get_print_or_pattern_type_display()
                if sku.print_or_pattern_type
                else "",
                # Occasion
                occasion,
                # Brand Color
                color,
                # Generic Keywords
                "",
            ]
        )

    # =========================================================
    # RESPONSE
    # =========================================================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="snapdeal_saree_template.xlsx"'
    )

    wb.save(response)

    return response


@login_required
def Myntra_Template1(request, sku_list):
    sku_list, _ = get_filtered_skus(request)

    validation_errors = validate_myntra_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/myntra_error.html",
            {"validation_errors": validation_errors},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sarees"

    # ======================================================
    # TOP HEADERS
    # ======================================================

    ws.merge_cells("A1:J1")
    ws["A1"] = "Version : 13"
    ws["A1"].alignment = Alignment(horizontal="left")

    yellow = PatternFill(fill_type="solid", start_color="FFFF99", end_color="FFFF99")
    pink = PatternFill(fill_type="solid", start_color="E53888", end_color="E53888")
    green = PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00")
    orange = PatternFill(fill_type="solid", start_color="EB5B00", end_color="EB5B00")

    ws.merge_cells("A2:AG2")
    ws["A2"] = (
        "Business (Information required for Style Creation/Legal Compliance/Order Tracking)"
    )
    ws["A2"].fill = green

    ws.merge_cells("AH2:BP2")
    ws["AH2"] = (
        "Discoverability - Attributes required for Product Description and Cataloguing"
    )
    ws["AH2"].fill = pink

    ws.merge_cells("BQ2:BU2")
    ws["BQ2"] = "Sizing - Mandatory Measurements"
    ws["BQ2"].fill = green

    ws.merge_cells("BV2:CB2")
    ws["BV2"] = "Images"
    ws["BV2"].fill = orange

    headers = [
        "styleId",
        "styleGroupId",
        "vendorSkuCode",
        "vendorArticleNumber",
        "vendorArticleName",
        "brand",
        "Manufacturer Name and Address with Pincode",
        "Packer Name and Address with Pincode",
        "Importer Name and Address with Pincode",
        "Country Of Origin",
        "Country Of Origin2",
        "Country Of Origin3",
        "Country Of Origin4",
        "Country Of Origin5",
        "articleType",
        "Brand Size",
        "Standard Size",
        "is Standard Size present on Label",
        "Brand Colour (Remarks)",
        "GTIN",
        "HSN",
        "SKUCode",
        "MRP",
        "ISP",
        "AgeGroup",
        "Prominent Colour",
        "Second Prominent Colour",
        "Third Prominent Colour",
        "FashionType",
        "Usage",
        "Year",
        "season",
        "AI Label",
        "List View Name",
        "Product Details",
        "styleNote",
        "materialCareDescription",
        "sizeAndFitDescription",
        "productDisplayName",
        "tags",
        "addedDate",
        "Color Variant GroupId",
        "Type",
        "Saree Fabric",
        "Blouse Fabric",
        "Blouse",
        "Pattern",
        "Print or Pattern Type",
        "Ornamentation",
        "Border",
        "Occasion",
        "Wash Care",
        "Trends",
        "Sustainable",
        "Main Trend",
        "Multipack Set",
        "Net Quantity Unit",
        "Theme",
        "Stitch",
        "Theme 1",
        "Technique",
        "Care for me",
        "Where-to-wear",
        "Style Tip",
        "BIS Expiry Date",
        "BIS Certificate Image URL",
        "BIS Certificate Number",
        "Net Quantity",
        "Bust ( Inches )",
        "Hip ( Inches )",
        "Outseam Length ( Inches )",
        "To Fit Waist ( Inches )",
        "Waist ( Inches )",
        "Front Image",
        "Side Image",
        "Back Image",
        "Detail Angle",
        "Look Shot Image",
        "Additional Image 1",
        "Additional Image 2",
    ]

    ws.append(headers)

    # ======================================================
    # STYLE GROUP
    # ======================================================

    sku_list = sku_list.order_by("van")

    previous_van = None
    style_group_id = 0

    for sku in sku_list:
        if previous_van != sku.van:
            style_group_id += 1
            previous_van = sku.van

        manufacturer = ""
        if sku.vendor:
            manufacturer = (
                f"{sku.vendor.company}, {sku.vendor.address}, {sku.vendor.pin}"
            )

        ws.append(
            [
                sku.old_sku or "",
                style_group_id,
                sku.sku or "",
                sku.van or "",
                sku.style_description or "",
                sku.brand.name if sku.brand else "",
                manufacturer,
                manufacturer,
                "",
                "India",
                "",
                "",
                "",
                "",
                sku.article_type.name.title() if sku.article_type else "",
                "One Size"
                if sku.size and sku.size.size == "Free Size"
                else (sku.size.size if sku.size else "FREE SIZE"),
                "Onesize"
                if sku.size and sku.size.size == "Free Size"
                else (sku.size.size if sku.size else "FREE SIZE"),
                "Yes",
                get_myntra_color(sku.color.color) if sku.color else "",
                "",
                sku.hsn or "5407",
                "",
                sku.mrp or "",
                sku.mrp or "",
                "Adults-Women",
                get_myntra_color(sku.color.color) if sku.color else "",
                "",
                "",
                "Fashion",
                "Ethnic",
                datetime.now().year if sku.style_description else "",
                "Spring",
                "",
                "",
                sku.style_description or "",
                sku.style_description or "",
                # content
                "",
                "",
                sku.style_description or "",
                "",
                "",
                "",
                # sku.get_type_display() if sku.type else "",
                get_myntra_technique(sku.type) if sku.type else "",
                get_myntra_saree_fabric(sku.saree_fabric) if sku.saree_fabric else "",
                get_myntra_blouse_fabric(sku.blouse_fabric)
                if sku.blouse_fabric
                else "",
                get_myntra_blouse(sku.blouse) if sku.blouse else "",
                get_myntra_pattern(sku.pattern) if sku.pattern else "",
                get_myntra_print_or_pattern_type(sku.print_or_pattern_type)
                if sku.print_or_pattern_type
                else "",
                get_myntra_ornamentation(sku.ornamentation)
                if sku.ornamentation
                else "",
                get_myntra_border(sku.border) if sku.border else "",
                # get_myntra_occasion(sku.get_occasion_display()) if sku.occasion else "",
                get_myntra_occasion(sku.occasion) if sku.occasion else "",
                "Dry Clean",
                "",
                "",
                "",
                "NA",
                "Piece",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "1",
                "",
                "",
                "",
                "",
                "",
                sku.product_image_link_1 or "",
                sku.product_image_link_2 or "",
                sku.product_image_link_3 or "",
                sku.product_image_link_4 or "",
                "",
                "",
                "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Myntra Listing Template.xlsx"'
    )

    wb.save(response)

    return response


@login_required
def View__VMS(request):
    if request.user.is_staff:
        vms_list = VMS.objects.all().order_by("-id")
    else:
        profile = Profile.objects.get(user=request.user)
        vms_list = VMS.objects.filter(vendor=profile).order_by("-id")

    if request.method == "GET":
        search_query = request.GET.get("search", "").strip()
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        # Filter by start date
        if start_date:
            try:
                parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d")
                vms_list = vms_list.filter(created_at__gte=parsed_start_date)
            except ValueError:
                pass  # Ignore invalid date

        # Filter by end date (inclusive of the entire day)
        if end_date:
            try:
                parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(
                    days=1
                )
                vms_list = vms_list.filter(created_at__lt=parsed_end_date)
            except ValueError:
                pass

        # Search filter by tracking_id
        if search_query:
            vms_list = vms_list.filter(Q(tracking_id__icontains=search_query))

        # Pagination
        paginator = Paginator(vms_list, 10)
        page = request.GET.get("page", 1)
        try:
            vms = paginator.page(page)
        except PageNotAnInteger:
            vms = paginator.page(1)
        except EmptyPage:
            vms = paginator.page(paginator.num_pages)

        context = {
            "vms": vms,
            "search_query": search_query,
            "start_date": start_date,
            "end_date": end_date,
        }

        return render(request, "vms/view_vms.html", context)


@login_required
def Delete__VMS(request):
    if request.method == "GET":
        search_query = request.GET.get("search", "").strip()
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        vms_list = VMS.objects.all().order_by("-id")

        # Filter by start date
        if start_date:
            try:
                parsed_start_date = datetime.strptime(start_date, "%Y-%m-%d")
                vms_list = vms_list.filter(created_at__gte=parsed_start_date)
            except ValueError:
                pass  # Ignore invalid date

        # Filter by end date (inclusive of the entire day)
        if end_date:
            try:
                parsed_end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(
                    days=1
                )
                vms_list = vms_list.filter(created_at__lt=parsed_end_date)
            except ValueError:
                pass

        # Search filter by tracking_id
        if search_query:
            vms_list = vms_list.filter(Q(tracking_id__icontains=search_query))

        # Pagination
        paginator = Paginator(vms_list, 10)
        page = request.GET.get("page", 1)
        try:
            vms = paginator.page(page)
        except PageNotAnInteger:
            vms = paginator.page(1)
        except EmptyPage:
            vms = paginator.page(paginator.num_pages)

        context = {
            "vms": vms,
            "search_query": search_query,
            "start_date": start_date,
            "end_date": end_date,
        }

        return render(request, "vms/view_vms.html", context)


@csrf_exempt
@login_required
def save_video(request):
    if request.method == "POST":
        tracking_id = request.POST.get("tracking_id")
        video_file = request.FILES.get("video_file")
        transaction_type = request.POST.get("transaction_type")
        profile = Profile.objects.get(user=request.user)

        vms = VMS.objects.create(
            vendor=profile,
            tracking_id=tracking_id,
            video_type=transaction_type,
            video_file=video_file,
        )

        return JsonResponse({"status": "success", "vms_id": vms.id})

    return JsonResponse({"status": "error"}, status=400)


@login_required
def record_video_page(request):
    return render(
        request, "vms/record_video.html"
    )  # use the full path inside 'templates'


@login_required
def download_database(request):
    db_path = settings.DATABASES["default"]["NAME"]
    db_file = Path(db_path)

    if not db_file.exists():
        return HttpResponse("Database file not found", status=404)

    # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamp = datetime.now(ZoneInfo("Asia/Kolkata"))
    filename = f"{timestamp}_database_backup.sqlite3"

    return FileResponse(
        open(db_file, "rb"),
        as_attachment=True,
        filename=filename,
    )


@login_required
def Meesho_Template1(request, sku_list):

    sku_list, _ = get_filtered_skus(request)

    validation_errors = validate_meesho_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/meesho_error.html",
            {"validation_errors": validation_errors},
        )

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Sarees-Fill this"

    # =========================================================
    # HEADERS
    # =========================================================

    headers = [
        "Fields + Description:",
        "ERROR STATUS",
        "ERROR MESSAGE",
        "Product Name",
        "Variation",
        "Meesho Price",
        "Wrong/Defective Returns Price",
        "MRP",
        "GST %",
        "HSN ID",
        "Net Weight (gms)",
        "Inventory",
        "Country of Origin",
        "Manufacturer Name",
        "Manufacturer Address",
        "Manufacturer Pincode",
        "Packer Name",
        "Packer Address",
        "Packer Pincode",
        "Importer Name",
        "Importer Address",
        "Importer Pincode",
        "Blouse",
        "Border",
        "Color",
        "Generic Name",
        "Net Quantity (N)",
        "Print or Pattern Type",
        "Saree Fabric",
        "Transparency",
        "Type",
        "Blouse Length Size",
        "Saree Length Size",
        "Image 1 (Front)",
        "Image 2",
        "Image 3",
        "Image 4",
        "Product ID / Style ID",
        "SKU ID",
        "Brand Name",
        "Group ID",
        "Product Description",
        "Blouse Color",
        "Blouse Fabric",
        "Blouse Pattern",
        "Border Width",
        "Brand",
        "Loom Type",
        "Occasion",
        "Ornamentation",
        "Pallu Details",
        "Pattern",
    ]

    ws.append(headers)

    # =========================================================
    # ROWS
    # =========================================================

    for sku in sku_list:
        color = get_marketplace_value(
            "MEESHO", "COLOR", sku.color.color if sku.color else None
        )

        blouse_color = get_marketplace_value(
            "MEESHO",
            "COLOR",
            sku.get_blouse_color_display() if sku.blouse_color else None,
        )

        blouse = get_marketplace_value(
            "MEESHO",
            "BLOUSE",
            sku.get_blouse_display() if sku.blouse else None,
        )

        border = get_marketplace_value(
            "MEESHO", "BORDER", sku.get_border_display() if sku.border else None
        )

        saree_fabric = get_marketplace_value(
            "MEESHO",
            "SAREE_FABRIC",
            sku.get_saree_fabric_display() if sku.saree_fabric else None,
        )

        blouse_fabric = get_marketplace_value(
            "MEESHO",
            "BLOUSE_FABRIC",
            sku.get_blouse_fabric_display() if sku.blouse_fabric else None,
        )

        occasion = get_marketplace_value(
            "MEESHO", "OCCASION", sku.get_occasion_display() if sku.occasion else None
        )

        ornamentation = get_marketplace_value(
            "MEESHO",
            "ORNAMENTATION",
            sku.get_ornamentation_display() if sku.ornamentation else None,
        )

        pattern = get_marketplace_value(
            "MEESHO", "PATTERN", sku.get_pattern_display() if sku.pattern else None
        )

        blouse_pattern = get_marketplace_value(
            "MEESHO",
            "BLOUSE_PATTERN",
            sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
        )

        print_pattern_type = get_marketplace_value(
            "MEESHO",
            "PRINT_OR_PATTERN_TYPE",
            sku.get_print_or_pattern_type_display()
            if sku.print_or_pattern_type
            else None,
        )

        technique = get_marketplace_value(
            "MEESHO",
            "TECHNIQUE",
            sku.get_type_display() if sku.type else None,
        )

        ws.append(
            [
                "",
                "",
                "",
                sku.style_description or "",
                sku.size.size if sku.size else "FREE SIZE",
                sku.sale_price or "",
                (sku.sale_price - Decimal("10.00") if sku.sale_price else ""),
                sku.mrp or "",
                5,
                5407,
                400,
                100,
                "India",
                sku.vendor.company if sku.vendor else "",
                sku.vendor.address if sku.vendor else "",
                sku.vendor.pin if sku.vendor else "",
                sku.vendor.company if sku.vendor else "",
                sku.vendor.address if sku.vendor else "",
                sku.vendor.pin if sku.vendor else "",
                "Not Required",
                "Not Required",
                "Not Required",
                blouse,
                border,
                color,
                sku.article_type.name if sku.article_type else "",
                "Single",
                print_pattern_type,
                saree_fabric,
                sku.get_transparency_display() if sku.transparency else "",
                technique,
                float(sku.blouse_length or 0.8),
                float(sku.saree_length or 5.5),
                sku.product_image_link_1 or "",
                sku.product_image_link_2 or "",
                sku.product_image_link_3 or "",
                sku.product_image_link_4 or "",
                sku.sku or "",
                sku.ref_no or "",
                sku.brand.name if sku.brand else "",
                "",
                sku.style_description or "",
                blouse_color,
                blouse_fabric,
                blouse_pattern,
                sku.get_border_width_display() if sku.border_width else "",
                sku.brand.name if sku.brand else "",
                sku.get_loom_type_display() if sku.loom_type else "",
                occasion,
                ornamentation,
                sku.get_pallu_details_display() if sku.pallu_details else "",
                pattern,
            ]
        )

    # =========================================================
    # RESPONSE
    # =========================================================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Sarees-10003-EXTERNAL-MeeshoTemplate2PricesGSTIN.xlsx"'
    )

    wb.save(response)

    return response


# @login_required
# def Flipkart_Template(request, sku_list):

#     sku_list, _ = get_filtered_skus(request)

#     validation_errors = validate_flipkart_template(sku_list)

#     if validation_errors:
#         return render(
#             request,
#             "validation/flipkart_error.html",
#             {"validation_errors": validation_errors},
#         )

#     wb = openpyxl.Workbook()

#     ws = wb.active
#     ws.title = "sari"

#     # =========================================================
#     # HEADERS
#     # =========================================================

#     headers = [
#         "Seller SKU ID",
#         "Group ID",
#         "Parent Variant FSN",
#         "Hidden",
#         "Listing Status",
#         "MRP (INR)",
#         "Your selling price (INR)",
#         "Fullfilment by",
#         "Procurement type",
#         "Procurement SLA (DAY)",
#         "Stock",
#         "Shipping provider",
#         "Local handling fee (INR)",
#         "Zonal handling fee (INR)",
#         "National handling fee (INR)",
#         "Length (CM)",
#         "Breadth (CM)",
#         "Height (CM)",
#         "Weight (KG)",
#         "HSN",
#         "Luxury Cess",
#         "Country Of Origin",
#         "Manufacturer Details",
#         "Packer Details",
#         "Importer Details",
#         "Tax Code",
#         "Minimum Order Quantity (MinOQ)",
#         "Brand",
#         "Occasion",
#         "Fabric",
#         "Pattern",
#         "Type",
#         "Sari Purity",
#         "Ideal For",
#         "Pack of",
#         "Fabric Care",
#         "Sari Length",
#         "Blouse Piece Length (m)",
#         "Sari Style",
#         "Brand Size",
#         "Brand Size - Measuring Unit",
#         "Style Code",
#         "Color",
#         "Brand Color",
#         "Blouse Piece Type",
#         "Main Image URL",
#         "Other Image URL 1",
#         "Other Image URL 2",
#         "Other Image URL 3",
#         "Other Image URL 4",
#         "Other Image URL 5",
#         "Other Image URL 6",
#         "Other Image URL 7",
#         "Main Palette Image URL",
#         "Pattern/Print Type",
#         "Border Details",
#         "Decorative Material",
#         "Blouse Fabric",
#         "Type of Embroidery",
#         "Video URL",
#         "Domestic Warranty",
#         "Domestic Warranty - Measuring Unit",
#         "Uniform",
#         "Transparent",
#         "Construction Type",
#         "Border Length",
#         "Blouse Pattern",
#         "Embroidery Method",
#         "EAN/UPC",
#         "EAN/UPC - Measuring Unit",
#         "Weight (kg)",
#         "Other Details",
#         "Description",
#         "Search Keywords",
#     ]

#     ws.append(headers)

#     # =========================================================
#     # ROWS
#     # =========================================================

#     for sku in sku_list:
#         color = get_marketplace_value(
#             "FLIPKART", "COLOR", sku.color.color if sku.color else None
#         )

#         blouse = get_marketplace_value(
#             "FLIPKART",
#             "BLOUSE",
#             sku.get_blouse_display() if sku.blouse else None,
#         )

#         border = get_marketplace_value(
#             "FLIPKART", "BORDER", sku.get_border_display() if sku.border else None
#         )

#         saree_fabric = get_marketplace_value(
#             "FLIPKART",
#             "SAREE_FABRIC",
#             sku.get_saree_fabric_display() if sku.saree_fabric else None,
#         )

#         blouse_fabric = get_marketplace_value(
#             "FLIPKART",
#             "BLOUSE_FABRIC",
#             sku.get_blouse_fabric_display() if sku.blouse_fabric else None,
#         )

#         occasion = get_marketplace_value(
#             "FLIPKART", "OCCASION", sku.get_occasion_display() if sku.occasion else None
#         )

#         ornamentation = get_marketplace_value(
#             "FLIPKART",
#             "ORNAMENTATION",
#             sku.get_ornamentation_display() if sku.ornamentation else None,
#         )

#         pattern = get_marketplace_value(
#             "FLIPKART", "PATTERN", sku.get_pattern_display() if sku.pattern else None
#         )

#         blouse_pattern = get_marketplace_value(
#             "FLIPKART",
#             "BLOUSE_PATTERN",
#             sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
#         )

#         print_pattern_type = get_marketplace_value(
#             "FLIPKART",
#             "PRINT_OR_PATTERN_TYPE",
#             sku.get_print_or_pattern_type_display()
#             if sku.print_or_pattern_type
#             else None,
#         )

#         technique = get_marketplace_value(
#             "MEESHO",
#             "TECHNIQUE",
#             sku.get_type_display() if sku.type else None,
#         )

#         border_width = get_marketplace_value(
#             "FLIPKART",
#             "BORDER_WIDTH",
#             sku.get_border_width_display() if sku.border_width else None,
#         )

#         ws.append(
#             [
#                 sku.sku or "",
#                 sku.id or "",
#                 "",
#                 "",
#                 "Active",
#                 sku.mrp or "",
#                 sku.sale_price + 200 or "",
#                 "Seller",
#                 "instock",
#                 "2",
#                 "100",
#                 "Flipkart",
#                 "0",
#                 "0",
#                 "0",
#                 "25",
#                 "20",
#                 "5",
#                 "0.400",
#                 "5407",
#                 "0",
#                 "India",
#                 ", ".join(
#                     filter(
#                         None,
#                         [
#                             sku.vendor.company if sku.vendor else "",
#                             sku.vendor.address if sku.vendor else "",
#                             str(sku.vendor.pin) if sku.vendor else "",
#                         ],
#                     )
#                 ),
#                 ", ".join(
#                     filter(
#                         None,
#                         [
#                             sku.vendor.company if sku.vendor else "",
#                             sku.vendor.address if sku.vendor else "",
#                             str(sku.vendor.pin) if sku.vendor else "",
#                         ],
#                     )
#                 ),
#                 "",
#                 "GST_5",
#                 "1",
#                 sku.brand.name if sku.brand else "",
#                 occasion,
#                 saree_fabric,
#                 pattern,
#                 technique,
#                 "Synthetic",
#                 "Women",
#                 "1",
#                 "Dry Clean Only",
#                 f"{float(sku.saree_length or 5.5):g}m",
#                 float(sku.blouse_length or 0.8),
#                 "Regular Sari",
#                 sku.size.size.split()[0].title() if sku.size else "Free",
#                 "Regular",
#                 sku.ref_no or "",
#                 color,
#                 color,
#                 blouse,
#                 sku.product_image_link_1 or "",
#                 sku.product_image_link_2 or "",
#                 sku.product_image_link_3 or "",
#                 sku.product_image_link_4 or "",
#                 "",
#                 "",
#                 "",
#                 "",
#                 "",
#                 print_pattern_type,
#                 border,
#                 ornamentation,
#                 blouse_fabric,
#                 "",
#                 "",
#                 "",
#                 "",
#                 "No",
#                 sku.get_transparency_display() if sku.transparency else "",
#                 sku.get_loom_type_display() if sku.loom_type else "",
#                 # sku.get_border_width_display() if sku.border_width else "",
#                 border_width,
#                 blouse_pattern,
#                 "Machine",
#                 "",
#                 "",
#                 "0.400",
#                 "",
#                 sku.style_description or "",
#                 "",
#                 # sku.vendor.company if sku.vendor else "",
#                 # sku.vendor.address if sku.vendor else "",
#                 # sku.vendor.pin if sku.vendor else "",
#                 # sku.vendor.company if sku.vendor else "",
#                 # sku.vendor.address if sku.vendor else "",
#                 # sku.vendor.pin if sku.vendor else "",
#             ]
#         )

#     # =========================================================
#     # RESPONSE
#     # =========================================================

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     part1 = uuid.uuid4().hex[:16]
#     part2 = f"{random.randint(1000, 9999)}-{random.randint(1000, 9999)}"
#     part3 = "".join(random.choices(string.ascii_uppercase + string.digits, k=14))

#     filename = f"C_sari_{part1}_{part2}_{part3}.xlsx"

#     response["Content-Disposition"] = f'attachment; filename="{filename}"'

#     wb.save(response)

#     return response


@login_required
def Flipkart_Template(request, sku_list):

    validation_errors = validate_flipkart_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/flipkart_error.html",
            {"validation_errors": validation_errors},
        )

    template_path = os.path.join(
        settings.BASE_DIR,
        "main",
        "marketplaces",
        "flipkart",
        "C_sari_f4d245f65fbb4767_0208-0936FK_REQQQKFLOKT6Q.xls",
    )

    rb = xlrd.open_workbook(
        template_path,
        formatting_info=True,
        on_demand=False,
    )

    sheet_index = rb.sheet_names().index("sari")

    wb = xl_copy(rb)

    ws = wb.get_sheet(sheet_index)

    # Excel G5
    row = 4
    col = 6

    # Process and write each SKU in the SAME loop
    for sku in sku_list:
        # ---------------------------------------------------------
        # Marketplace values for THIS SKU
        # ---------------------------------------------------------

        color = get_marketplace_value(
            "FLIPKART",
            "COLOR",
            sku.color.color if sku.color else None,
        )

        blouse = get_marketplace_value(
            "FLIPKART",
            "BLOUSE",
            sku.get_blouse_display() if sku.blouse else None,
        )

        border = get_marketplace_value(
            "FLIPKART",
            "BORDER",
            sku.get_border_display() if sku.border else None,
        )

        saree_fabric = get_marketplace_value(
            "FLIPKART",
            "SAREE_FABRIC",
            sku.get_saree_fabric_display() if sku.saree_fabric else None,
        )

        blouse_fabric = get_marketplace_value(
            "FLIPKART",
            "BLOUSE_FABRIC",
            sku.get_blouse_fabric_display() if sku.blouse_fabric else None,
        )

        occasion = get_marketplace_value(
            "FLIPKART",
            "OCCASION",
            sku.get_occasion_display() if sku.occasion else None,
        )

        ornamentation = get_marketplace_value(
            "FLIPKART",
            "ORNAMENTATION",
            sku.get_ornamentation_display() if sku.ornamentation else None,
        )

        pattern = get_marketplace_value(
            "FLIPKART",
            "PATTERN",
            sku.get_pattern_display() if sku.pattern else None,
        )

        blouse_pattern = get_marketplace_value(
            "FLIPKART",
            "BLOUSE_PATTERN",
            sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
        )

        print_pattern_type = get_marketplace_value(
            "FLIPKART",
            "PRINT_OR_PATTERN_TYPE",
            sku.get_print_or_pattern_type_display()
            if sku.print_or_pattern_type
            else None,
        )

        technique = get_marketplace_value(
            "FLIPKART",
            "TECHNIQUE",
            sku.get_type_display() if sku.type else None,
        )

        border_width = get_marketplace_value(
            "FLIPKART",
            "BORDER_WIDTH",
            sku.get_border_width_display() if sku.border_width else None,
        )

        # ---------------------------------------------------------
        # Build values for THIS SKU
        # ---------------------------------------------------------

        values = [
            sku.sku or "",
            sku.id or "",
            "",
            "",
            "Active",
            sku.mrp or "",
            sku.sale_price + 200 if sku.sale_price else "",
            "Seller",
            "instock",
            "2",
            "100",
            "Flipkart",
            "0",
            "0",
            "0",
            "25",
            "20",
            "5",
            "0.400",
            "5407",
            "0",
            "India",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            "",
            "GST_5",
            "1",
            sku.brand.name if sku.brand else "",
            occasion,
            saree_fabric,
            pattern,
            technique,
            "Synthetic",
            "Women",
            "1",
            "Dry Clean Only",
            f"{float(sku.saree_length or 5.5):g}m",
            float(sku.blouse_length or 0.8),
            "Regular Sari",
            sku.size.size.split()[0].title() if sku.size else "Free",
            "Regular",
            sku.sku or "",
            # COLOR
            color,
            color,
            blouse,
            sku.product_image_link_1 or "",
            sku.product_image_link_2 or "",
            sku.product_image_link_3 or "",
            sku.product_image_link_4 or "",
            "",
            "",
            "",
            "",
            "",
            print_pattern_type,
            border,
            ornamentation,
            blouse_fabric,
            "",
            "",
            "",
            "",
            "No",
            sku.get_transparency_display() if sku.transparency else "",
            sku.get_loom_type_display() if sku.loom_type else "",
            border_width,
            blouse_pattern,
            "Machine",
            "",
            "",
            "0.400",
            "",
            sku.style_description or "",
            "::".join(
                word.strip()
                for word in re.split(
                    r"[,\s]+",
                    sku.key_words or "",
                )
                if word.strip()
            ),
        ]

        # ---------------------------------------------------------
        # Write THIS SKU to Excel
        # ---------------------------------------------------------

        for offset, value in enumerate(values):
            ws.write(
                row,
                col + offset,
                value,
            )

        # Move to next Excel row
        row += 1

    # -------------------------------------------------------------
    # Response
    # -------------------------------------------------------------

    response = HttpResponse(content_type="application/vnd.ms-excel")

    filename = "C_sari_f4d245f65fbb4767_0208-0936FK_REQQQKFLOKT6Q.xls"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


@login_required
def Snapdeal_Template(request, sku_list):

    validation_errors = validate_snapdeal_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/snapdeal_error.html",
            {"validation_errors": validation_errors},
        )

    template_path = os.path.join(
        settings.BASE_DIR,
        "main",
        "marketplaces",
        "snapdeal",
        "Sdb673_Women's Saree_1727_1787240709759.xlsx",
    )

    wb = load_workbook(template_path)

    # IMPORTANT: select correct sheet
    ws = wb["Women's Saree_1727"]

    # Remove merged cells from data area
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 5:
            ws.unmerge_cells(str(merged))

    # Start writing from A5
    row = 3
    col = 1

    for sku in sku_list:
        color = get_marketplace_value(
            "SNAPDEAL", "COLOR", sku.color.color if sku.color else None
        )

        blouse = get_marketplace_value(
            "SNAPDEAL", "BLOUSE", sku.get_blouse_display() if sku.blouse else None
        )

        border = get_marketplace_value(
            "SNAPDEAL", "BORDER", sku.get_border_display() if sku.border else None
        )

        saree_fabric = get_marketplace_value(
            "SNAPDEAL",
            "SAREE_FABRIC",
            sku.get_saree_fabric_display() if sku.saree_fabric else None,
        )

        blouse_fabric = get_marketplace_value(
            "SNAPDEAL",
            "BLOUSE_FABRIC",
            sku.get_blouse_fabric_display() if sku.blouse_fabric else None,
        )

        occasion = get_marketplace_value(
            "SNAPDEAL", "OCCASION", sku.get_occasion_display() if sku.occasion else None
        )

        ornamentation = get_marketplace_value(
            "SNAPDEAL",
            "ORNAMENTATION",
            sku.get_ornamentation_display() if sku.ornamentation else None,
        )

        pattern = get_marketplace_value(
            "SNAPDEAL", "PATTERN", sku.get_pattern_display() if sku.pattern else None
        )

        blouse_pattern = get_marketplace_value(
            "SNAPDEAL",
            "BLOUSE_PATTERN",
            sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
        )

        print_pattern_type = get_marketplace_value(
            "SNAPDEAL",
            "PRINT_OR_PATTERN_TYPE",
            sku.get_print_or_pattern_type_display()
            if sku.print_or_pattern_type
            else None,
        )

        technique = get_marketplace_value(
            "SNAPDEAL", "TECHNIQUE", sku.get_type_display() if sku.type else None
        )

        border_width = get_marketplace_value(
            "SNAPDEAL",
            "BORDER_WIDTH",
            sku.get_border_width_display() if sku.border_width else None,
        )

        values = [
            sku.id or "",
            sku.sku or "",
            sku.brand.name if sku.brand else "",
            "",
            color,
            saree_fabric,
            blouse,
            technique,
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            float(sku.saree_length or 5.5),
            float(sku.blouse_length or 0.8),
            "India",
            "Not Applicable",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            "1",
            pattern,
            "Pack of 1",
            print_pattern_type,
            "Regular Saree",
            sku.ref_no or "",
            sku.mrp or "",
            (sku.sale_price + 300) if sku.sale_price else "",
            "100",
            "2",
            "5",
            "20",
            "25",
            "0.400",
            sku.product_image_link_1 or "",
            sku.product_image_link_2 or "",
            sku.product_image_link_3 or "",
            sku.product_image_link_4 or "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            sku.style_description or "",
            "",
            "",
            blouse_fabric,
            color,  # blouse color
            border,
            "1.08",
            "0.400",
            "Saree",
            "",
            blouse_pattern,
            print_pattern_type,
            # keep remaining values exactly same as your existing list
        ]

        # Write starting from A5
        for index, value in enumerate(values):
            ws.cell(row=row, column=col + index, value=value)

        row += 1

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    template_filename = os.path.basename(template_path)

    filename = template_filename

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def Myntra_Template(request, sku_list):
    validation_errors = validate_myntra_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/myntra_error.html",
            {"validation_errors": validation_errors},
        )

    template_path = os.path.join(
        settings.BASE_DIR,
        "main",
        "marketplaces",
        "myntra",
        "Myntra-Sku-Template-2026-08-03.xlsx",
    )

    wb = load_workbook(template_path)

    ws = wb["Sarees"]

    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 5:
            ws.unmerge_cells(str(merged))

    row = 3
    col = 1

    for sku in sku_list:
        # -------------------------------
        # Myntra mappings per SKU
        # -------------------------------

        color = get_marketplace_value(
            "MYNTRA",
            "COLOR",
            sku.color.color if sku.color else None,
        )

        blouse = get_marketplace_value(
            "MYNTRA",
            "BLOUSE",
            sku.get_blouse_display() if sku.blouse else None,
        )

        border = get_marketplace_value(
            "MYNTRA",
            "BORDER",
            sku.get_border_display() if sku.border else None,
        )

        saree_fabric = get_marketplace_value(
            "MYNTRA",
            "SAREE_FABRIC",
            sku.get_saree_fabric_display() if sku.saree_fabric else None,
        )

        blouse_fabric = get_marketplace_value(
            "MYNTRA",
            "BLOUSE_FABRIC",
            sku.get_blouse_fabric_display() if sku.blouse else None,
        )

        occasion = get_marketplace_value(
            "MYNTRA",
            "OCCASION",
            sku.get_occasion_display() if sku.occasion else None,
        )

        ornamentation = get_marketplace_value(
            "MYNTRA",
            "ORNAMENTATION",
            sku.get_ornamentation_display() if sku.ornamentation else None,
        )

        pattern = get_marketplace_value(
            "MYNTRA",
            "PATTERN",
            sku.get_pattern_display() if sku.pattern else None,
        )

        blouse_pattern = get_marketplace_value(
            "MYNTRA",
            "BLOUSE_PATTERN",
            sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
        )

        print_pattern_type = get_marketplace_value(
            "MYNTRA",
            "PRINT_OR_PATTERN_TYPE",
            sku.get_print_or_pattern_type_display()
            if sku.print_or_pattern_type
            else None,
        )

        technique = get_marketplace_value(
            "MYNTRA",
            "TECHNIQUE",
            sku.get_type_display() if sku.type else None,
        )

        border_width = get_marketplace_value(
            "MYNTRA",
            "BORDER_WIDTH",
            sku.get_border_width_display() if sku.border_width else None,
        )

        # -------------------------------
        # Excel row data
        # -------------------------------

        size = (
            "Onesize"
            if sku.size and sku.size.size == "Free Size"
            else (sku.size.size if sku.size else "Onesize")
        )

        values = [
            sku.id or "",
            sku.sku or "",
            sku.ref_no or "",
            sku.sku or "",
            sku.brand.name if sku.brand else "",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            "",
            "India",
            "",
            "",
            "",
            "",
            "Sarees",
            size,
            size,
            "Yes",
            color,
            "",
            "54075490",
            "",
            sku.mrp or "",
            sku.mrp or "",
            "Adults-Women",
            color,
            "",
            "",
            "Fashion",
            occasion,
            "2026",
            "Spring",
            "",
            "",
            sku.style_description or "",
            "",
            "Dry Clean Only",
            "",
            "",
            "",
            "",
            "",
            technique,
            saree_fabric,
            blouse_fabric,
            blouse,
            pattern,
            print_pattern_type,
            ornamentation,
            border,
            occasion,
            "Dry Clean",
            "",
            "Regular",
            "",
            "NA",
            "Piece",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "1",
            "",
            "",
            "",
            "",
            "",
            sku.product_image_link_1 or "",
            sku.product_image_link_2 or "",
            sku.product_image_link_3 or "",
            sku.product_image_link_4 or "",
            "",
            "",
            "",
        ]

        # Write row
        for offset, value in enumerate(values):
            ws.cell(row=row + 1, column=col + offset + 1, value=value)

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = "myntra_template.xlsx"

    date_str = datetime.now().strftime("%Y-%m-%d")
    new_filename = f"Myntra-Sku-Template-{date_str}{os.path.splitext(filename)[1]}"

    response["Content-Disposition"] = f'attachment; filename="{new_filename}"'

    wb.save(response)

    return response


@login_required
def Meesho_Template(request, sku_list):

    validation_errors = validate_meesho_template(sku_list)

    if validation_errors:
        return render(
            request,
            "validation/meesho_error.html",
            {"validation_errors": validation_errors},
        )

    template_path = os.path.join(
        settings.BASE_DIR,
        "main",
        "marketplaces",
        "mee",
        "Sarees-10003-EXTERNAL-MeeshoTemplate2PricesGSTIN.xlsx",
    )

    wb = load_workbook(template_path)

    ws = wb["Sarees-Fill this"]

    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 5:
            ws.unmerge_cells(str(merged))

    row = 4
    col = 3

    for sku in sku_list:
        # -------------------------------
        # Myntra mappings per SKU
        # -------------------------------

        color = get_marketplace_value(
            "MEESHO",
            "COLOR",
            sku.color.color if sku.color else None,
        )

        blouse = get_marketplace_value(
            "MEESHO",
            "BLOUSE",
            sku.get_blouse_display() if sku.blouse else None,
        )

        border = get_marketplace_value(
            "MEESHO",
            "BORDER",
            sku.get_border_display() if sku.border else None,
        )

        saree_fabric = get_marketplace_value(
            "MEESHO",
            "SAREE_FABRIC",
            sku.get_saree_fabric_display() if sku.saree_fabric else None,
        )

        blouse_fabric = get_marketplace_value(
            "MEESHO",
            "BLOUSE_FABRIC",
            sku.get_blouse_fabric_display() if sku.blouse else None,
        )

        occasion = get_marketplace_value(
            "MEESHO",
            "OCCASION",
            sku.get_occasion_display() if sku.occasion else None,
        )

        ornamentation = get_marketplace_value(
            "MEESHO",
            "ORNAMENTATION",
            sku.get_ornamentation_display() if sku.ornamentation else None,
        )

        pattern = get_marketplace_value(
            "MEESHO",
            "PATTERN",
            sku.get_pattern_display() if sku.pattern else None,
        )

        blouse_pattern = get_marketplace_value(
            "MEESHO",
            "BLOUSE_PATTERN",
            sku.get_blouse_pattern_display() if sku.blouse_pattern else None,
        )

        print_pattern_type = get_marketplace_value(
            "MEESHO",
            "PRINT_OR_PATTERN_TYPE",
            sku.get_print_or_pattern_type_display()
            if sku.print_or_pattern_type
            else None,
        )

        technique = get_marketplace_value(
            "MEESHO",
            "TECHNIQUE",
            sku.get_type_display() if sku.type else None,
        )

        border_width = get_marketplace_value(
            "MEESHO",
            "BORDER_LENGTH",
            sku.get_border_width_display() if sku.border_width else None,
        )

        # -------------------------------
        # Excel row data
        # -------------------------------

        size = (
            "Onesize"
            if sku.size and sku.size.size == "Free Size"
            else (sku.size.size if sku.size else "Onesize")
        )

        values = [
            str(sku.article_type) if sku.article_type else "",
            str(sku.size) if sku.size else "",
            sku.sale_price or "",
            (sku.sale_price - 10) if sku.sale_price is not None else "",
            sku.mrp or "",
            "5",
            "5407",
            "400",
            "500",
            "India",
            sku.vendor.company if sku.vendor else "",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            str(sku.vendor.pin) if sku.vendor else "",
            sku.vendor.company if sku.vendor else "",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            str(sku.vendor.pin) if sku.vendor else "",
            sku.vendor.company if sku.vendor else "",
            ", ".join(
                filter(
                    None,
                    [
                        sku.vendor.company if sku.vendor else "",
                        sku.vendor.address if sku.vendor else "",
                        str(sku.vendor.pin) if sku.vendor else "",
                    ],
                )
            ),
            str(sku.vendor.pin) if sku.vendor else "",
            blouse,
            border,
            color,
            str(sku.article_type) if sku.article_type else "",
            "Single",
            print_pattern_type,
            saree_fabric,
            "No" if str(sku.transparency).lower() == "no" else "Yes",
            technique,
            float(sku.blouse_length or 0.8),
            float(sku.saree_length or 5.5),
            sku.product_image_link_1 or "",
            sku.product_image_link_2 or "",
            sku.product_image_link_3 or "",
            sku.product_image_link_4 or "",
            sku.sku or "",
            sku.sku or "",
            sku.brand.name if sku.brand else "",
            sku.id or "",
            sku.style_description or "",
            color,
            blouse_fabric,
            blouse_pattern,
            border_width,
            sku.brand.name if sku.brand else "",
            sku.loom_type or "",
            occasion,
            ornamentation,
            sku.pallu_details,
            pattern,
        ]

        # Write row
        for offset, value in enumerate(values):
            ws.cell(row=row + 1, column=col + offset + 1, value=value)

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = "myntra_template.xlsx"

    date_str = datetime.now().strftime("%Y-%m-%d")
    new_filename = f"Myntra-Sku-Template-{date_str}{os.path.splitext(filename)[1]}"

    response["Content-Disposition"] = f'attachment; filename="{new_filename}"'

    wb.save(response)

    return response

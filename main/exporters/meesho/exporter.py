import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side

from .constants import MEESHO_HEADERS
from .mapper import MeeshoMapper
from .validator import MeeshoValidator
from ...models import SKU


class MeeshoExporter:

    def export(self, user):

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sarees-Fill this"

        # -------------------------
        # ROW 1: TITLE
        # -------------------------
        
        ws["A1"] = "Sarees Template (Women Fashion/Ethnic Wear/Sarees, Blouses & Petticoats/Sarees)"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        ws["A1"].font = Font(size=18, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws["E1"] = "Sarees Template (Women Fashion/Ethnic Wear/Sarees, Blouses & Petticoats/Sarees)"
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=len(MEESHO_HEADERS))

        ws["E1"].font = Font(size=9)
        ws["E1"].alignment = Alignment(horizontal="left", vertical="center")

        # -------------------------
        # ROW 2: FIELD RULES
        # -------------------------
        row2 = [
            "Field Names",
            "Do not fill these 2 columns. To be filled by Meesho only.",
            "Do not fill these 2 columns. To be filled by Meesho only."
        ]

        compulsory = "* Compulsory Field"
        optional = "Optional Field"

        remaining = len(MEESHO_HEADERS) - len(row2)

        for i in range(remaining):
            row2.append(compulsory if i < 20 else optional)

        ws.append(row2)

        for col in range(1, len(MEESHO_HEADERS) + 1):
            ws.cell(row=2, column=col).font = Font(size=9)

        # -------------------------
        # ROW 3: LONG DESCRIPTIONS
        # -------------------------
        MEESHO_DESCRIPTIONS = [
            "Fields + Description:",
            "ERROR STATUS\n\nFor system use, don't modify",
            "ERROR MESSAGE\n\nFor system use, don't modify",
            "Product Name\n\nPlease enter product name...",
            "Variation\n\nSize selection rules...",
            "Meesho Price\n\nRegular selling price...",
            "Wrong/Defective Returns Price\n\nReturn rules...",
            "MRP\n\nMaximum Retail Price...",
            "GST %\n\nProduct GST percent",
            "HSN ID\n\nSelect HSN from dropdown",
            "Net Weight (gms)\n\nProduct weight excluding packaging",
            "Inventory\n\nStock available",
            "Country of Origin\n\nManufacturing country",
            "Manufacturer Name\n\nManufacturer details",
            "Manufacturer Address\n\nFull address required",
            "Manufacturer Pincode\n\nPincode only",
            "Packer Name\n\nPacking entity",
            "Packer Address\n\nPacking address",
            "Packer Pincode\n\nPacking pincode",
            "Importer Name\n\nImporter details",
            "Importer Address\n\nImporter address",
            "Importer Pincode\n\nImporter pincode",
            "Blouse\n\nSelect blouse option",
            "Border\n\nBorder selection",
            "Color\n\nProduct color",
            "Generic Name\n\nCommon product name",
            "Net Quantity (N)\n\nNumber of units",
            "Print or Pattern Type\n\nPattern selection",
            "Saree Fabric\n\nFabric type",
            "Transparency\n\nTransparency level",
            "Type\n\nProduct type",
            "Blouse Length Size\n\nMeter unit",
            "Saree Length Size\n\nMeter unit",
            "Image 1 (Front)\n\nUpload link required",
            "Image 2\n\nUpload link required",
            "Image 3\n\nUpload link required",
            "Image 4\n\nUpload link required",
            "Product ID / Style ID\n\nUnique style code",
            "SKU ID\n\nSeller SKU identifier",
            "Brand Name\n\nBrand information",
            "Group ID\n\nGrouping identifier",
            "Product Description\n\nShort product description",
            "Blouse Color\n\nBlouse color selection",
            "Blouse Fabric\n\nBlouse fabric type",
            "Blouse Pattern\n\nPattern type",
            "Border Width\n\nBorder width",
            "Brand\n\nBrand selection",
            "Loom Type\n\nLoom category",
            "Occasion\n\nUsage occasion",
            "Ornamentation\n\nDecoration type",
            "Pallu Details\n\nPallu description",
            "Pattern\n\nPattern type"
        ]

        # format spacing
        MEESHO_DESCRIPTIONS = [
            d if i == 0 else f"\n\n{d}\n"
            for i, d in enumerate(MEESHO_DESCRIPTIONS)
        ]

        ws.append(MEESHO_DESCRIPTIONS)

        for col in range(1, len(MEESHO_HEADERS) + 1):
            cell = ws.cell(row=4, column=col)
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # -------------------------
        # ROW 4 EXTRA CELL
        # -------------------------
        ws["A4"] = "Tutorial Link"
        ws["A4"].font = Font(size=12, bold=False)
        ws["A4"].alignment = Alignment(horizontal="left",vertical="center")

        ws["G4"] = "Watch Explainer Video"
        ws["G4"].font = Font(size=11, bold=False)
        ws["G4"].alignment = Alignment(horizontal="left",vertical="center")

        ws["AH4"] = "Watch Explainer Video"
        ws["AH4"].font = Font(size=11, bold=False)
        ws["AH4"].alignment = Alignment(horizontal="left",vertical="center")

        ws["AM4"] = "Watch Explainer Video"
        ws["AM4"].font = Font(size=11, bold=False)
        ws["AM4"].alignment = Alignment(horizontal="left",vertical="center")

        # -------------------------
        # BORDER A2 -> AZ4
        # -------------------------
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # AZ = 52 columns
        for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=52):
            for cell in row:
                cell.border = thin_border

        # -------------------------
        # DATA ROWS
        # -------------------------
        skus = SKU.objects.select_related(
            "brand", "vendor", "color", "size", "article_type"
        )

        profile = getattr(user, "profile", None)
        validator = MeeshoValidator()

        row_idx = 5

        for sku in skus:
            errors = validator.validate(sku)
            if errors:
                continue

            row = MeeshoMapper(sku, profile).map()
            ws.append(row)

            for col_idx in range(1, len(row) + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(size=9)

            row_idx += 1

        # -------------------------
        # RESPONSE
        # -------------------------
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="meesho.xlsx"'

        wb.save(response)
        return response